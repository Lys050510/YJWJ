"""
模拟开箱图片批量处理工具
功能：裁黑边 → 统一尺寸(128x128) → 锐化 → 转WebP → 重命名 → 更新xlsx → 删原图
生成错误登记表：assets/box/_处理错误报告.xlsx
"""
import os, sys, glob, re, json
from datetime import datetime
from PIL import Image, ImageFilter, ImageOps

# ============ 配置 ============
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BOX_DIR = os.path.join(ROOT, 'assets', 'box')

# Item icons
ICON_SIZE = 128          # 统一输出尺寸
ICON_PADDING = 4         # 内容距画布边缘的最小间距
DARK_THRESHOLD = 25      # RGB 各通道低于此值视为黑边（0-255）
SHARPEN_RADIUS = 0.5     # 锐化半径
SHARPEN_AMOUNT = 80      # 锐化强度 (0-200%)
WEBP_QUALITY_ICON = 90   # 图标 WebP 质量

# Overview images
OVERVIEW_MAX = 800       # 概览图最长边
WEBP_QUALITY_OVERVIEW = 85

# ============ 工具函数 ============
def sanitize_filename(name):
    """清理文件名中的非法字符"""
    # 保留中文、字母、数字、·、-、_
    result = []
    illegal = '<>:"/\\|?*'
    for ch in name:
        if ch in illegal:
            continue
        result.append(ch)
    return ''.join(result).strip()

def trim_dark_border(im):
    """检测并裁切黑边（四边扫描，找到首列/行非暗色像素的位置）"""
    w, h = im.size
    pixels = im.load()

    # 左边界 → 扫描列，找到第一列有亮色像素的 x
    left = 0
    for x in range(w):
        has_light = False
        for y in range(h):
            r, g, b = pixels[x, y][:3]
            if r > DARK_THRESHOLD or g > DARK_THRESHOLD or b > DARK_THRESHOLD:
                has_light = True
                break
        if has_light:
            left = x
            break

    # 右边界
    right = w - 1
    for x in range(w - 1, -1, -1):
        has_light = False
        for y in range(h):
            r, g, b = pixels[x, y][:3]
            if r > DARK_THRESHOLD or g > DARK_THRESHOLD or b > DARK_THRESHOLD:
                has_light = True
                break
        if has_light:
            right = x
            break

    # 上边界
    top = 0
    for y in range(h):
        has_light = False
        for x in range(w):
            r, g, b = pixels[x, y][:3]
            if r > DARK_THRESHOLD or g > DARK_THRESHOLD or b > DARK_THRESHOLD:
                has_light = True
                break
        if has_light:
            top = y
            break

    # 下边界
    bottom = h - 1
    for y in range(h - 1, -1, -1):
        has_light = False
        for x in range(w):
            r, g, b = pixels[x, y][:3]
            if r > DARK_THRESHOLD or g > DARK_THRESHOLD or b > DARK_THRESHOLD:
                has_light = True
                break
        if has_light:
            bottom = y
            break

    # 确保有效裁切
    if left >= right or top >= bottom:
        return im  # 全部是暗色，不裁

    cropped = im.crop((left, top, right + 1, bottom + 1))
    return cropped

def process_item_icon(filepath, output_path):
    """处理单个物品图标：裁黑边 → 缩放到128x128居中 → 锐化 → WebP"""
    try:
        im = Image.open(filepath)
        if im.mode not in ('RGBA', 'RGB'):
            im = im.convert('RGBA')

        orig_size = im.size

        # 1. 裁黑边
        im = trim_dark_border(im)
        cropped_size = im.size

        # 2. 缩放到画布内（留 padding）
        canvas_w = ICON_SIZE - ICON_PADDING * 2
        canvas_h = ICON_SIZE - ICON_PADDING * 2
        cw, ch = im.size

        ratio = min(canvas_w / cw, canvas_h / ch)
        new_w = max(1, int(cw * ratio))
        new_h = max(1, int(ch * ratio))
        im = im.resize((new_w, new_h), Image.LANCZOS)

        # 3. 居中放到透明画布
        canvas = Image.new('RGBA', (ICON_SIZE, ICON_SIZE), (0, 0, 0, 0))
        offset_x = (ICON_SIZE - new_w) // 2
        offset_y = (ICON_SIZE - new_h) // 2
        canvas.paste(im, (offset_x, offset_y), im if im.mode == 'RGBA' else None)

        # 4. 锐化
        canvas = canvas.filter(ImageFilter.UnsharpMask(
            radius=SHARPEN_RADIUS, percent=SHARPEN_AMOUNT
        ))

        # 5. 保存 WebP
        canvas.save(output_path, 'webp', quality=WEBP_QUALITY_ICON, method=6)

        return {
            'success': True,
            'orig_size': orig_size,
            'cropped_size': cropped_size,
            'output_size': (ICON_SIZE, ICON_SIZE)
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

def process_overview(filepath, output_path):
    """处理概览图：缩放到最长边800px → WebP"""
    try:
        im = Image.open(filepath)
        if im.mode in ('RGBA', 'P', 'PA'):
            im = im.convert('RGBA')
        else:
            im = im.convert('RGB')

        w, h = im.size
        longest = max(w, h)

        if longest > OVERVIEW_MAX:
            ratio = OVERVIEW_MAX / longest
            new_w = int(w * ratio)
            new_h = int(h * ratio)
            im = im.resize((new_w, new_h), Image.LANCZOS)

        save_kwargs = {'quality': WEBP_QUALITY_OVERVIEW, 'method': 6}
        if im.mode == 'RGBA':
            save_kwargs['lossless'] = False

        im.save(output_path, 'webp', **save_kwargs)
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}

# 设置 stdout 编码为 UTF-8（Windows 兼容）
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ============ 主流程 ============
def main():
    import openpyxl

    errors = []  # 错误登记
    stats = {'processed_items': 0, 'processed_overviews': 0, 'deleted_pngs': 0, 'failed': 0}

    # 遍历所有箱子目录
    box_dirs = sorted([
        d for d in os.listdir(BOX_DIR)
        if os.path.isdir(os.path.join(BOX_DIR, d)) and not d.startswith('_')
    ])

    print(f'{"="*60}')
    print(f'模拟开箱图片批量处理')
    print(f'目录数: {len(box_dirs)}')
    print(f'{"="*60}\n')

    for box_name in box_dirs:
        box_path = os.path.join(BOX_DIR, box_name)
        xlsx_path = os.path.join(box_path, '箱子概况.xlsx')

        if not os.path.exists(xlsx_path):
            errors.append({
                '目录': box_name, '物品名': '-', '错误类型': '缺少xlsx',
                '详情': '箱子概况.xlsx 不存在'
            })
            continue

        print(f'📦 {box_name} ...')

        # 读取 xlsx
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        overview_filename = str(ws.cell(1, 3).value).strip() if ws.cell(1, 3).value else ''

        # 读取物品列表 (row >= 3, 有名称和文件名)
        items = []
        for row_idx in range(3, ws.max_row + 1):
            name = str(ws.cell(row_idx, 1).value).strip() if ws.cell(row_idx, 1).value else ''
            quality = str(ws.cell(row_idx, 2).value).strip() if ws.cell(row_idx, 2).value else ''
            old_filename = str(ws.cell(row_idx, 3).value).strip() if ws.cell(row_idx, 3).value else ''

            if not name and not old_filename:
                continue  # 空白行/备注行
            if not old_filename:
                errors.append({
                    '目录': box_name, '物品名': name, '错误类型': 'xlsx缺少文件名',
                    '详情': f'第{row_idx}行："{name}"没有对应图片文件名'
                })
                continue

            items.append({
                'row': row_idx,
                'name': name,
                'quality': quality,
                'old_filename': old_filename
            })

        # 处理概览图
        overview_old_path = None
        if overview_filename:
            overview_old_path = os.path.join(box_path, overview_filename)
            if os.path.exists(overview_old_path):
                overview_new_path = os.path.join(box_path, '概览.webp')
                result = process_overview(overview_old_path, overview_new_path)
                if result['success']:
                    # 更新 xlsx C1
                    ws.cell(1, 3).value = '概览.webp'
                    stats['processed_overviews'] += 1
                    print(f'   ✅ 概览: {overview_filename} → 概览.webp')
                else:
                    errors.append({
                        '目录': box_name, '物品名': '【概览图】', '错误类型': '处理失败',
                        '详情': f'{overview_filename}: {result["error"]}'
                    })
                    stats['failed'] += 1
            else:
                errors.append({
                    '目录': box_name, '物品名': '【概览图】', '错误类型': '文件缺失',
                    '详情': f'xlsx引用 {overview_filename} 但文件不存在'
                })

        # 处理物品图标
        for idx, item in enumerate(items, start=1):
            old_path = os.path.join(box_path, item['old_filename'])
            new_filename = f'{idx:02d}.webp'
            new_path = os.path.join(box_path, new_filename)

            if not os.path.exists(old_path):
                errors.append({
                    '目录': box_name, '物品名': item['name'], '错误类型': '文件缺失',
                    '详情': f'第{item["row"]}行引用 {item["old_filename"]} 但文件不存在'
                })
                stats['failed'] += 1
                continue

            result = process_item_icon(old_path, new_path)
            if result['success']:
                # 更新 xlsx
                ws.cell(item['row'], 3).value = new_filename
                stats['processed_items'] += 1
                orig = result['orig_size']
                print(f'   ✅ [{idx:02d}] {item["name"]}: {orig[0]}x{orig[1]} → {new_filename}')
            else:
                errors.append({
                    '目录': box_name, '物品名': item['name'], '错误类型': '处理失败',
                    '详情': f'{item["old_filename"]}: {result["error"]}'
                })
                stats['failed'] += 1

        # 保存 xlsx
        wb.save(xlsx_path)
        wb.close()

        # 删除原始 PNG 文件
        png_files = glob.glob(os.path.join(box_path, '*.png'))
        for png in png_files:
            try:
                os.remove(png)
                stats['deleted_pngs'] += 1
            except Exception as e:
                errors.append({
                    '目录': box_name, '物品名': '-', '错误类型': '删除失败',
                    '详情': f'{os.path.basename(png)}: {e}'
                })

        print(f'   🗑️ 已删除 {len(png_files)} 个原 PNG 文件')
        print()

    # ============ 生成错误报告 ============
    error_path = os.path.join(BOX_DIR, '_处理错误报告.xlsx')
    if errors:
        ew = openpyxl.Workbook()
        ews = ew.active
        ews.title = '错误登记'
        ews.append(['序号', '目录', '物品名称', '错误类型', '详情'])
        for i, err in enumerate(errors, start=1):
            ews.append([i, err['目录'], err['物品名'], err['错误类型'], err['详情']])
        # 调整列宽
        ews.column_dimensions['A'].width = 6
        ews.column_dimensions['B'].width = 20
        ews.column_dimensions['C'].width = 30
        ews.column_dimensions['D'].width = 14
        ews.column_dimensions['E'].width = 60
        ew.save(error_path)
        print(f'\n⚠️ 错误登记表: {error_path} ({len(errors)} 条)')
    else:
        print(f'\n✅ 全部处理成功，无错误！')

    # ============ 汇总 ============
    print(f'\n{"="*60}')
    print(f'处理完成！')
    print(f'  物品图标: {stats["processed_items"]} 个')
    print(f'  概览图:   {stats["processed_overviews"]} 张')
    print(f'  删除PNG:  {stats["deleted_pngs"]} 个')
    print(f'  失败:     {stats["failed"]} 个')
    print(f'  错误条目: {len(errors)} 条')
    print(f'{"="*60}')

if __name__ == '__main__':
    main()
